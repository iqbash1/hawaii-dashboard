#!/usr/bin/env python3
"""
backfill-food-insecurity-ers.py

Backfills food_insecurity_rate for all three-year periods (2006-2008 through 2022-2024)
using the USDA ERS Food Security interactive charts data file.

Source: https://www.ers.usda.gov/media/649/data-file-for-interactive-charts.xlsx
Sheet:  "Food security by State"

Columns:
  Year                         - three-year period (e.g. "2006-2008")
  State                        - two-letter state abbreviation
  Food insecurity prevalence   - percent food insecure (e.g. 12.2 means 12.2%)

Values are stored in state-data.js as decimal fractions (0.122 for 12.2%).

Usage:
    python scripts/backfill-food-insecurity-ers.py
    python scripts/backfill-food-insecurity-ers.py --dry-run
"""

import io
import json
import os
import sys
import argparse
import urllib.request
import ssl as _ssl_for_bypass
_ssl_for_bypass._create_default_https_context = _ssl_for_bypass._create_unverified_context  # local-env workaround for self-signed cert in TLS chain

import openpyxl

REPO_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
STATE_DATA_PATH = os.path.join(REPO_ROOT, 'js', 'state-data.js')

ERS_URL = 'https://www.ers.usda.gov/media/649/data-file-for-interactive-charts.xlsx'
ERS_SHEET = 'Food security by State'

# Mapping from 2-letter abbreviation to state-data.js state name
ABBR_TO_STATE = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawai\u02BBi', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming',
}

# Skip DC and national total
SKIP_ABBRS = {'DC', 'U.S. total', 'U.S.'}


def load_state_data():
    with open(STATE_DATA_PATH, 'r', encoding='utf-8') as f:
        raw = f.read()
    marker = 'const STATE_DATA ='
    idx = raw.find(marker)
    body = raw[idx + len(marker):].strip().rstrip(';').strip()
    return json.loads(body)


def write_state_data(data):
    js = 'const STATE_DATA = ' + json.dumps(data, indent=2, ensure_ascii=False) + ';\n'
    with open(STATE_DATA_PATH, 'w', encoding='utf-8') as f:
        f.write(js)


def fetch_ers_excel():
    print(f'Fetching USDA ERS food security Excel file...')
    req = urllib.request.Request(ERS_URL, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=60) as resp:
        raw = resp.read()
    print(f'  Downloaded {len(raw):,} bytes')
    return raw


def parse_ers_excel(raw_bytes):
    """Parse ERS Excel file, return {period: {state_name: decimal_rate}}"""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)

    if ERS_SHEET not in wb.sheetnames:
        raise ValueError(f'Sheet "{ERS_SHEET}" not found. Sheets: {wb.sheetnames}')

    ws = wb[ERS_SHEET]
    data = {}
    header = None

    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if header is None:
            # Row 1 is a merged title; row 2 has the actual column headers
            if not any(cell is not None for cell in row):
                continue
            raw_header = [str(c).strip() if c is not None else '' for c in row]
            # Check if this looks like a real header row (must have both 'Year' and 'State' as distinct values)
            if not (any(h == 'Year' for h in raw_header) and any(h == 'State' for h in raw_header)):
                continue  # skip title rows
            header = raw_header
            # Find column indices
            try:
                year_col = next(i for i, h in enumerate(header) if 'Year' in h or 'year' in h)
                state_col = next(i for i, h in enumerate(header) if 'State' in h or 'state' in h)
                prev_col = next(i for i, h in enumerate(header) if 'insecurity prevalence' in h.lower() and 'margin' not in h.lower())
            except StopIteration:
                print(f'  Header: {header}')
                raise ValueError('Could not find required columns in header')
            continue

        if not any(cell is not None for cell in row):
            continue

        # Normalize en-dash (–) to hyphen (-) in period strings like "2006–2008"
        period = str(row[year_col]).strip().replace('\u2013', '-') if row[year_col] is not None else ''
        abbr = str(row[state_col]).strip() if row[state_col] is not None else ''
        prev_raw = row[prev_col]

        # Skip national total and DC
        if abbr in SKIP_ABBRS or abbr.startswith('U.S'):
            continue

        # Map abbreviation to state name
        state_name = ABBR_TO_STATE.get(abbr)
        if not state_name:
            continue  # Skip unknown abbreviations

        if prev_raw is None:
            continue

        try:
            prev_pct = float(prev_raw)
        except (TypeError, ValueError):
            continue

        # Convert percent to decimal (12.2 -> 0.122)
        decimal_val = round(prev_pct / 100, 4)

        if period not in data:
            data[period] = {}
        data[period][state_name] = decimal_val

    wb.close()
    return data


def main():
    parser = argparse.ArgumentParser(
        description='Backfill food_insecurity_rate from USDA ERS Excel into state-data.js'
    )
    parser.add_argument('--dry-run', action='store_true', help='Fetch and parse but do not write')
    parser.add_argument('--no-skip-existing', action='store_true',
                        help='Overwrite periods already in state-data.js')
    args = parser.parse_args()

    raw = fetch_ers_excel()
    data = parse_ers_excel(raw)

    periods = sorted(data.keys())
    print(f'  Parsed {len(periods)} periods: {periods[0]} to {periods[-1]}')

    # Coverage check
    for period in periods[:3]:
        n = len(data[period])
        hi = data[period].get('Hawai\u02BBi', 'n/a')
        print(f'  {period}: {n} states, HI={hi}')
    if len(periods) > 3:
        period = periods[-1]
        n = len(data[period])
        hi = data[period].get('Hawai\u02BBi', 'n/a')
        print(f'  ...{period}: {n} states, HI={hi}')

    if args.dry_run:
        print('\n[dry-run] No changes written.')
        sys.exit(0)

    state_data = load_state_data()
    metric = state_data['food_insecurity_rate']
    existing_periods = set(metric['data'].keys())
    skip = not args.no_skip_existing
    added, skipped = 0, 0

    for period, states in data.items():
        if skip and period in existing_periods:
            skipped += 1
        else:
            metric['data'][period] = states
            added += 1

    # Sort periods chronologically (they start with year like "2006-2008")
    metric['data'] = dict(sorted(metric['data'].items()))
    all_periods = list(metric['data'].keys())
    hi_first = metric['data'].get(all_periods[0], {}).get('Hawai\u02BBi', 'n/a')
    hi_last = metric['data'].get(all_periods[-1], {}).get('Hawai\u02BBi', 'n/a')

    print(f'\n  food_insecurity_rate: +{added} added, {skipped} skipped')
    print(f'  Now covers {all_periods[0]} to {all_periods[-1]} ({len(all_periods)} periods)')
    print(f'  HI earliest={hi_first}, latest={hi_last}')

    write_state_data(state_data)
    print('Done.')


if __name__ == '__main__':
    main()
